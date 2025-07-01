import io
import os
import json
import shutil
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from typing import Literal, Optional, Union, List


from tools.views import *
from langchain.tools import tool
from tools.utils import col_to_colidx



# DATAFRAME_REGISTRY = {
#     'my_dataframe': {
#         'dataframe': pd.DataFrame(),  # Placeholder for DataFrame object
#         'file_path': '/path/to/file.xlsx',
#         'sheet_name': 'Sheet1',
#         'origin_header_row': 0
#     }
# }
DATAFRAME_REGISTRY = {}

@tool('Done Tool', args_schema=Done)
def done_tool(answer: str):
    """
    A tool to indicate that a task is completed.
    """
    return answer

@tool('Human Tool', args_schema=HumanTool)
def human_tool(question: str):
    """
    A tool for human input.
    """
    if not question:
        return "No question provided for human input."
    answer = input(f"Human Input Required: {question}\nYour answer: ")

    return f"Human answer: {answer}"

@tool('Load DataFrame Tool', args_schema=LoadDataFrame)
def load_dataframe_tool(df_name: str, file_path: str, sheet_name: Optional[Union[str, int]] = 0, origin_header_row: int = 0):
    df = pd.read_excel(file_path, sheet_name=sheet_name, header=origin_header_row)
    DATAFRAME_REGISTRY[df_name] = {
        'dataframe': df, # DataFrame object
        'file_path': file_path, 
        'sheet_name': sheet_name,
        'origin_header_row': origin_header_row
    }
    return f"DataFrame Object '{df_name}' Registered from file '{file_path}' with sheet '{sheet_name}', file path '{file_path}' and header row {origin_header_row}."

@tool('Excel Head Tool', args_schema=ExcelHead)
def excel_head_tool(file_path: str, head: int):
    """
    A tool to get the first few rows of an Excel file.
    Returns the first few rows as a  Matrix format string
    """
    df = pd.read_excel(file_path, header=None)
    df_head = df.head(head).fillna("")
    result = [f"Row {idx+1}: {list(row)}" for idx, row in zip(df_head.index, df_head.values)]
    return f"The first {head} rows (including empty rows) from file '{file_path}':\n" + "\n".join(result)

@tool('Excel Info Tool', args_schema=ExcelInfo)
def excel_info_tool(df_name: str):
    """
    A tool to get information about an Excel file.
    """
    df_info = DATAFRAME_REGISTRY.get(df_name)
    if not df_info:
        return f"DataFrame Object '{df_name}' not found."

    file_path = df_info['file_path']
    sheet_name = df_info['sheet_name']
    origin_header_row = df_info['origin_header_row']

    buffer = io.StringIO()
    info = df_info['dataframe'].info(buf=buffer)
    info_output = buffer.getvalue()
    buffer.close()
    info_str = f"DataFrame Object '{df_name}' Info:\n"
    info_str += f"File Path: {file_path}\n"
    info_str += f"Sheet Name: {sheet_name}\n"
    info_str += f"Header Row: {origin_header_row}\n"
    info_str += f"Info of DataFrame:\n{info_output}\n"
    return info_str

@tool('Read DataFrame Tool', args_schema=ReadDataFrame)
def read_dataframe_tool(df_name: str, row: Optional[Union[int, List[int]]] = None, col: Optional[Union[int, str, List[Union[int, str]]]] = None):
    if df_name not in DATAFRAME_REGISTRY:
        return f"DataFrame Object '{df_name}' not found."
    df = DATAFRAME_REGISTRY[df_name]['dataframe']
    
    # 处理行参数
    rows = None
    if row is not None:
        if isinstance(row, int):
            if row < 0 or row >= len(df):
                return f"Row index {row} out of range for DataFrame '{df_name}'."
            rows = [row]
        elif isinstance(row, list):
            if any(r < 0 or r >= len(df) for r in row):
                return f"Row indices {row} out of range for DataFrame '{df_name}'."
            rows = row
        else:
            return "Invalid 'row' parameter type; must be int or list of ints."
        
    # 处理列参数
    cols = None
    if col is not None:
        if isinstance(col, (int, str)):
            idx = col_to_colidx(df, col)
            if idx == -1:
                return f"Column '{col}' not found in DataFrame '{df_name}'."
            cols = [idx]
        elif isinstance(col, list):
            cols = []
            for c in col:
                idx = col_to_colidx(df, c)
                if idx == -1:
                    return f"Column '{c}' not found in DataFrame '{df_name}'."
                cols.append(idx)
        else:
            return "Invalid 'col' parameter type; must be int, str or list of these."


    try:
        df_selected = df

        if cols is not None:
            df_selected = df_selected.iloc[:, cols]
        
        if rows is not None:
            df_selected = df_selected.iloc[rows, :]

        matrix = df_selected.fillna("").values.tolist() # 将DataFrame转换为二维列表

        return f"Reading DataFrame '{df_name}':\n{matrix}"
    except Exception as e:
        return f"Error reading DataFrame '{df_name}': {str(e)}"
    
    
    
@tool('Write DataFrame Tool', args_schema=WriteDataFrame)
def write_dataframe_tool(df_name: str,
                         start_row: int = 0,
                         start_col: Optional[Union[int, str]] = 0,
                         step: int = 1,
                         values: Union[Any, List[Any], List[List[Any]]] = None,
                         axis: Literal['row', 'column', 'matrix'] = None):
    
    if values is None:
        return "No value provided for writing."

    if axis not in ['row', 'column', 'matrix']:
        return f"Invalid axis '{axis}'. Must be one of 'row', 'column', or 'matrix'."
    

    if df_name not in DATAFRAME_REGISTRY:
        return f"DataFrame Object '{df_name}' not found."
    df = DATAFRAME_REGISTRY[df_name]['dataframe']
    max_row, max_col = df.shape

    if start_col is None:
        col_idx = 0
    else:
        col_idx = col_to_colidx(df, start_col) if isinstance(start_col, str) else start_col
    if not isinstance(col_idx, int) or col_idx < 0 or col_idx >= max_col:
        return f"Invalid start_col: '{start_col}' resolved to index {col_idx}, which is out of range."
    
    if start_row < 0 or start_row >= max_row:
        return f"Start row {start_row} is out of range."
    
    # 统一 values 为二维列表
    if not isinstance(values, list):
        values = [[values]] if axis == "matrix" else [values]
    elif all(not isinstance(v, list) for v in values):
        if axis == 'row':
            values = [values]
        elif axis == 'column':
            values = [[v] for v in values]

    try:
        for i, row_vals in enumerate(values):
            for j, val in enumerate(row_vals):
                if axis == "row":
                    r = start_row + i * step
                    c = col_idx + j * step
                elif axis == "column":
                    r = start_row + j * step
                    c = col_idx + i * step
                else:  # matrix
                    r = start_row + i * step
                    c = col_idx + j * step
                if r >= max_row or c >= max_col:
                    return f"Writing out of bounds: row {r} or column {c} exceeds DataFrame dimensions ({max_row}, {max_col})."
                # 类型兼容处理，避免 FutureWarning
                if df.dtypes.iloc[c] != 'object':
                    df[df.columns[c]] = df[df.columns[c]].astype('object')
                df.iat[r, c] = val
    except Exception as e:
        return f"Error writing to DataFrame '{df_name}': {str(e)}"
        
    DATAFRAME_REGISTRY[df_name]['dataframe'] = df
    return f"DataFrame '{df_name}' updated with {len(values)} values, rows starting at row {start_row}, cols starting at col {start_col}, step {step}, axis '{axis}'."

@tool('DataFrame to Excel Tool', args_schema=DataFrame2Excel)
def dataframe2excel_tool(df_name: str):
    """
    A tool to convert a DataFrame to an Excel file.
    """
    if df_name not in DATAFRAME_REGISTRY:
        return f"DataFrame Object '{df_name}' not found."
    df_info = DATAFRAME_REGISTRY[df_name]
    df = df_info['dataframe']
    file_path = df_info['file_path']
    sheet_name = df_info['sheet_name']
    origin_header_row = df_info['origin_header_row']

    # Backup the original file if required
    base, ext = os.path.splitext(file_path)
    backup_path = base + datetime.now().strftime(".backup_%Y%m%d_%H%M%S") + ext
    try:
        shutil.copyfile(file_path, backup_path)
    except Exception as e:
        return f"Failed to backup the original file: {e}"
    
    
    try:
        wb = load_workbook(backup_path)
        if isinstance(sheet_name, int):
            ws = wb.worksheets[sheet_name]
        else:
            ws = wb[sheet_name]
            
        # Write headers
        for col_idx, col_name in enumerate(df.columns, start=1):
            cell = ws.cell(row=origin_header_row + 1, column=col_idx)
            if isinstance(cell, MergedCell):
                continue
            cell.value = col_name

        # Write data
        for row_idx, row in enumerate(df.itertuples(index=False), start=origin_header_row + 2):
            for col_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if isinstance(cell, MergedCell):
                    continue
                cell.value = value

        wb.save(backup_path)

    except Exception as e:
        return f"Failed to write DataFrame to Excel file: {e}"

    return f"DataFrame '{df_name}' written to Excel file '{backup_path}' in sheet '{sheet_name}' with header row {origin_header_row}."



